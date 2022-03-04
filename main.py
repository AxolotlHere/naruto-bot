import discord
import discord.ext.commands as commands
import time
import openpyxl

bot = commands.Bot(command_prefix="n!")
bot.remove_command('help')

workbook_1 = openpyxl.load_workbook(r"userdata.xlsx")
workbook_2 = openpyxl.load_workbook(r"overall_data.xlsx")
user_data = workbook_1["Sheet1"]


@bot.command(name="help")
async def help(ctx):
    help_embed = discord.Embed(color=0x00FFFF)
    help_embed.title = "BOT-HELP"
    help_embed.url = "https://youtu.be/dQw4w9WgXcQ"
    help_embed.description = 'prefix : `n!`\ntype `n!start` to create a new profile'
    help_embed.set_author(name=ctx.author.display_name, icon_url=ctx.author.avatar_url)
    help_embed.add_field(name='Field 1', value='text1 comes here', inline=False)
    help_embed.add_field(name='Field 2', value='text2 comes here', inline=False)
    help_embed.add_field(name='Field 3', value='text3 comes here', inline=False)
    help_embed.set_footer(text=f'Time - {time.strftime("%H:%M:%S")}')

    await ctx.send(embed=help_embed)


@bot.event
async def on_ready():
    for i in bot.guilds:
        for j in i.text_channels:
            if str(j) == "✇🍥-da-ttebayo-🍥✇":
                embed = discord.Embed(title='BOT-Restart',
                                      url='https://youtu.be/dQw4w9WgXcQ',
                                      description='The BOT has successfully restarted',
                                      colour=0xFF0000)
                await j.send(embed=embed)
                break
        else:
            await i.create_text_channel("✇🍥-da-ttebayo-🍥✇")
            for j in i.text_channels:
                if str(j) == "✇🍥-da-ttebayo-🍥✇":
                    embed = discord.Embed(title='BOT-Restart',
                                          url='https://youtu.be/dQw4w9WgXcQ',
                                          description='The BOT has successfully restarted',
                                          colour=0xFF0000)
                    await j.send(embed=embed)
                    break


@bot.command(name="terminate")
async def terminate(ctx):
    if ctx.author.id in [898395530349133825, 904730228000194591, 818002642789203989, 915143047329771530]:
        embed = discord.Embed(color=0xFF0000)
        embed.title = 'Terminating the Bot'
        embed.description = f'Kakashi Chidori\'d me on the request of `{ctx.author.display_name}`'
        embed.set_footer(text=f'Terminated at {time.strftime("%H:%M:%S")}')
        embed.set_image(
            url='https://cdn.discordapp.com/attachments/947401413795913729/947445493980622888/images_-_2022-02-27T161754.539.jpeg')
        await ctx.channel.send(embed=embed)
        await bot.close()
        print('Logging out', f'Bot is being terminated by `{ctx.author.display_name}`',
              f'Terminated at {time.strftime("%H:%M:%S")}', sep='\n')
        quit()


@bot.command(name='start')
async def start(ctx):
    user_exist = False
    for i in range(2, user_data.max_row + 1):
        print(str(user_data.cell(row=i, column=1).value))
        print(str(ctx.message.author))
        if str(ctx.message.author) == str(user_data.cell(row=i, column=1).value):

            user_exist = True
            break

    if user_exist==True:
        embed = discord.Embed(color=0xcdabef)
        embed.set_image(url="https://cdn.discordapp.com/attachments/947401562886668348/949015002629677137/efa7824401f311a04ed4df82d2032bfac086188de5cd89a529b589cb2977d3a7_1.jpg")
        embed.title = "My foolish little brother"
        embed.description = "User already exist...It is certain that I\'m dumb but that is far from being dumb as you .....just kidding :)"
        embed.set_footer(text="User " + str(ctx.author.name) + " already exist.")
    else:
        max_row_const = user_data.max_row + 1
        user_data.cell(row=max_row_const, column=1).value = str(ctx.message.author)
        user_data.cell(row=max_row_const, column=2).value = str([])
        user_data.cell(row=max_row_const, column=3).value = str([])
        user_data.cell(row=max_row_const, column=4).value = str([])
        user_data.cell(row=max_row_const, column=5).value = 0
        workbook_1.save(filename=r"userdata.xlsx")
        embed = discord.Embed(color=0xcdabef)
        embed.title = f'Welcome to the Shinobi world {ctx.author.display_name}'
        embed.set_image(url='https://cdn.discordapp.com/attachments/947830664424751185/947831654494384128/Z.png')
    await ctx.channel.send(embed=embed)


token = "OTI2ODAyMjE3MzkzMDg2NDg0.YdA90w.X6e8s-GO58mnXF3aXyikiuGGmdI"
bot.run(token)
